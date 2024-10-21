from xls2xlsx import XLS2XLSX
from flask import Flask, jsonify, request, send_file
from werkzeug.utils import secure_filename
from flask_cors import CORS
from io import BytesIO
from openpyxl import Workbook, load_workbook
import os
import json

app = Flask(__name__) 
CORS(app)


ALLOWED_EXTENSIONS = set(['xls', 'xlsx' ])
UPLOAD_FOLDER = os.path.abspath(os.path.join(os.path.dirname(__file__), 'Downloads'))
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 500 * 1000 * 1000  # 500 MB
app.config['CORS_HEADER'] = 'application/json'

def allowedFile(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
           

splcourses = {"AI":["CSET225", "Intelligent Model Design using AI"], "Cloud Computing":["CSET232", "Design of Cloud Architectural Solutions"] }
maincourses = {"CSET208":"Ethics for Engineers, Patents, Copyrights and IPR","CSET302":"Automata Theory and Computability",\
    "CSET305":"High Performance Computing","CSET303":"Seminar on Special Topics in Emerging Areas", "CSET304":"Competitive Programming",}
splelectives = {"Natural Language Processing":"CSET346", "Blockchain Technologies: Platforms & Applications":"CSET350", "Penetration Testing, Auditing and Ethical Testing": "CSET363", "Time Series Analysis": "CSET369", "Big Data Analytics and Business Intelligence": "CSET371", "Augmented Reality": "CSET337", "Build and Release Management in DevOps": "CSET448", "Front-End Web UI Frameworks and Tools: Bootstrap" : "CSET457", "Cloud Infrastructure and Services": "CSET358", "Quantum Computing for Data Analysis": "CSET474", "IoT Analytics": "CSET480", "Augmented Reality and ARCore":"CSET425", "Design and Manufacturing for Digital Products": "CSET433", "User Centered Design": "CSET321" }
electives = {"Soft Computing": "CSET326", "Compiler Construction" : "CSET323", "Software Project Management" : "CSET3240", "Engineering Optimization" : "CSET329"}
      
           
def parse(specialisation,elective, splelec,  wb):
    elecCourse = electives[elective]
    splcourse = splcourses[specialisation]
    splecourse = splelectives[splelec]
    ttwb = wb
    tt = ttwb.active
    
    classes = []
    c = 0
    for i in range(2,7):
        for j in range(5,14):
            
            match (i-2):
                case 0:
                    dstart = '12'
                case 1:
                    dstart = '6'
                case 2:
                    dstart = '7'
                case 3:
                    dstart = '8'
                case 4:
                    dstart = '9'
            match (j-5):
                case 0:
                    tstart = '08'
                case 1:
                    tstart = '09'
                case 2:
                    tstart = '10'
                case 3:
                    tstart = '11'
                case 4:
                    tstart = '12'
                case 5:
                    tstart = '13'
                case 6:
                    tstart = '14'
                case 7:
                    tstart = '15'
                case 8:
                    tstart = '16'
            value = tt.cell(row = j, column = i).value
            
            if value == None or value == "":
                continue
            k = next((k for k in maincourses.keys() if k in value), None)
            if k is not None:
                i1 = value.index("{")
                i2 = value.index("}")
                room = value[i1+1:i2]
                name = maincourses[k]
            elif splcourse[0] in value:
                i1 = value.index(splcourse[0])
                value = value[i1:]
                i2 = value.index("}")
                value = value[:i2+1]
                print("This is value",value)
                i3 = value.index("{")
                room = value[i3+1:i2]
                name = splcourse[1]
            elif elecCourse in value:
                i1 = value.index(elecCourse)
                value = value[i1:]
                i2 = value.index("}")
                value = value[:i2+1]
                i3 = value.index("{")
                room = value[i3+1:i2]
                name = elective
            elif splecourse in value:
                i1 = value.index(splecourse)
                value = value[i1:]
                i2 = value.index("}")
                value = value[:i2+1]
                i3 = value.index("{")
                room = value[i3+1:i2]
                name = splelec
            else:
                continue
            if "(L)" in value:
                name += " Lecture"
            elif "(T)" in value:
                name += " Tutorial"
            elif "(P)" in value:
                name += " Lab"
            classes.append({"course": value, "room": room, "name": name, "day": dstart, "time": tstart})
        c += 1
    return classes
           
           
           
           
@app.route('/upload', methods=['POST', 'GET'])
def fileUpload():
    if request.method == 'POST':
        file = request.files.getlist('files')
        spl = request.form['spl']
        spl_elective = request.form['spl_elective']
        elective = request.form['elective']
        print( "PAY ATTENTION ", spl, elective, spl_elective)
        filename = ""
        print(request.files, "....")
        for f in file:
            print(f.filename)
            filename = secure_filename(f.filename)
            print(allowedFile(filename))
            if allowedFile(filename):
                f.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                if filename.split('.')[1] == 'xlsx':
                    wb = load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    classes = parse(spl , elective, spl_elective,  wb)
                else:
                    x2x = XLS2XLSX(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    wb = x2x.to_xlsx()
                    classes = parse(spl , elective, spl_elective,  wb)
                
            else:
                return jsonify({'message': 'File type not allowed'}), 400
            

        return jsonify(json.dumps({'class':classes}))
        # return jsonify({"name": filename, "status": "success"})
    else:
        return jsonify({"status": "Upload API GET Request Running"})



# on the terminal type: curl http://127.0.0.1:5000/ 
# returns hello world when we use GET. 
# returns the data that we send when we use POST. 
@app.route('/', methods = ['GET', 'POST']) 
def home(): 
	if(request.method == 'GET'): 
		data = "hello world"
		return jsonify({'data': data}) 


if __name__ == '__main__': 
	app.run(debug = True) 